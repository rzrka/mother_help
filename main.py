import pandas as pd
import openpyxl

TYPE_CLOTHES = {('руб', 'рубаш', 'рубашка'):
                  {('д/р', 'дл.рук',):
                       {('повседневаная', 'повс', 'повсед', 'голубая', 'голуб.', 'гол.'):
                            'Рубашка с длинным рукавом повседневная голубого цвета (2 шт. в соответствии с приказом)',
                       ('парад.', 'пардная', 'пар.'):
                            'Рубашка с длинным рукавом парадная белого цвета'
                       },
                  ('к/р', 'кор.рук', 'кор.т.с.', 'кор.с.р', 'кор.р'):
                       {('повседневаная', 'повс', 'повсед', 'голубая', 'голуб.', 'гол.'):
                            'Рубашка с коротким рукавом повседневная голубого цвета  (2 шт. в соответствии с приказом)',
                       ('парад.', 'пардная', 'пар.'):
                            'Рубашка с коротким рукавом парадная белого цвета'
                       },
                   ('поло',): 'Рубашка поло',
                   },
                  ('брюки',):
                      {('лет',):
                           'Брюки летние',
                       ('дем'):
                            'Брюки демисезонные',
                      },
                  ('Джемпер',): 'Джемпер трикотажный',
                  ('куртка',):
                    {
                        ('вет',): 'Куртка ветровка',
                        ('зим',): 'Куртка зимняя'
                    }

              }
NAMES = {
    # ('Брюки муж.лет.',): 'Брюки летние мужские',
    ('Убор головной зимний мужской', 'Шапка'): 'Мужской головной убор зимний (трикотажная шапка)',
    # ('Куртка мужская зимняя',): 'Куртка мужская зимняя',
    # ('Руб.муж. д/р лог."МГТ" парад.',
     # 'Рубашка муж. д/р парад.'): 'Рубашка мужская с длинным рукавом парадная белого цвета',
    # ('Рубашка повс.муж.д/р',
     # 'Рубашка повс.муж.дл.рук.'): 'Рубашка мужская с длинным рукавом повседневная голубого цвета (2 шт. в соответствии с приказом)',
    # (
        # 'Рубашка повс.муж.к/р',): 'Рубашка мужская с коротким рукавом повседневная голубого цвета  (2 шт. в соответствии с приказом)',
    # ('Брюки муж.демис.',): 'Брюки демисезонные мужские',
    # ('Куртка ветр.муж.',): 'Куртка ветровка мужская',
    # ('Рубашка поло',): 'Рубашка поло мужская',
    # ('Джемпер трик. мужской', 'Джемпер муж'): 'Джемпер трикотажный мужской ',
    ('Жилет муж.утепл.',): 'Жилет утепленный мужской',
    # ('Рубашка мужская парадная к/р',
     # 'Рубашка муж. к/р парад.',): 'Рубашка мужская с коротким рукавом парадная белого цвета',
}


def search_clothes(cloth, iter_clothes=TYPE_CLOTHES):
    for type_cloth in iter_clothes:
        for el in type_cloth:
            if el.lower() in cloth.lower():
                iter_clothes = iter_clothes[type_cloth]
                if type(iter_clothes) is str:
                    return iter_clothes
                else:
                    return search_clothes(cloth, iter_clothes)

def clothes(tabs):
    '''
    Получение данных от бухгалтерии
    :param tabs:
    :return:
    '''
    df = pd.read_excel('data/excel.xlsx',
                       skiprows=9)
    tab_id = df['Unnamed: 2']
    skiped = df['Unnamed: 1']
    cloths = df['Unnamed: 11']
    dates = df['Unnamed: 10']
    counts = df['Unnamed: 18']
    data = {}
    for skip, tab_el, cloth, date, count in zip(skiped, tab_id, cloths, dates, counts):

        if skip == '*' or tab_el not in tabs:
            continue
        else:
            if not data.get(tab_el):
                data[tab_el] = {}
            re_cloth = search_clothes(cloth)
            if re_cloth:
                if data[tab_el].get(re_cloth):
                    data[tab_el][re_cloth]['count'] += 1
                else:
                    data[tab_el][re_cloth] = {'date': date.strftime('%d-%m-%Y'), 'count': count}
    return data

def set_data(data):
    '''
    Создание excel по данным от бугалтерии
    :param data:
    :return:
    '''
    cells = {
        'Брюки летние мужские': (20,),
        'Мужской головной убор зимний (трикотажная шапка)': (16,),
        'Куртка мужская зимняя': [15],
        'Рубашка мужская с длинным рукавом парадная белого цвета': [10],
        'Рубашка мужская с длинным рукавом повседневная голубого цвета (2 шт. в соответствии с приказом)': (8, 9),
        'Рубашка мужская с коротким рукавом повседневная голубого цвета  (2 шт. в соответствии с приказом)': (17, 18),
        'Брюки демисезонные мужские': [11],
        'Куртка ветровка мужская': [14],
        'Рубашка поло мужская': [21],
        'Джемпер трикотажный мужской ': [12],
        'Жилет утепленный мужской': [13],
        'Рубашка мужская с коротким рукавом парадная белого цвета': [19]
    }
    wb_obj = openpyxl.load_workbook('data/Образец матрицы заполнения.xlsx')
    sheet_obj = wb_obj.active

    i = 6

    for tab in data:
        for j in (range(8, 22 + 1)):
            cell_obj = sheet_obj.cell(row=i, column=j)
            cell_obj.value = 'не выдано'
        tab_id = sheet_obj.cell(row=i, column=1)
        tab_id.value = tab
        for cloth in data[tab]:
            date = data[tab][cloth]['date']
            count = data[tab][cloth]['count']
            if count == 2.0:
                for num_coll in cells[cloth]:
                    cell_cloth = sheet_obj.cell(row=i, column=num_coll)
                    cell_cloth.value = date
            else:
                cell_cloth = sheet_obj.cell(row=i, column=cells[cloth][0])
                cell_cloth.value = date
        i += 1
    wb_obj.save('data/ПрограммаКО.xlsx')


def get_tab_id():
    '''
    Получение всех нужных табельных номеров
    :return:
        list
    '''
    wb_obj = openpyxl.load_workbook('data/все водители с телефонами 14.09.2022.xlsx')
    sheet_obj = wb_obj.active
    start = 2
    tabs = set()
    for tab_num in range(start, sheet_obj.max_row + 1):
        tabs.add(int(sheet_obj.cell(row=tab_num, column=1).value))

    return tabs


tabs = get_tab_id()
cloths = clothes(tabs)
set_data(cloths)


#TODO
'''
1. не корректное чтение файла бугалтерии
'''
