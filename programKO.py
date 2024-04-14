import pandas as pd
import openpyxl
from add_size import ManSizes, WomanSizes
import os
import re

FILENAME = 'data/Корпоративная одежда ФЗ.xlsx'

TYPE_CLOTHES_EXCLUDED = 'перчатки|сиг|Полукомб|Галстук|Фуражка|Рукавицы|Ботинки|Пиджак'.lower()

TYPE_CLOTHES = {('поло',): ['Рубашка поло'],
                ('руб',):
                  {
                       ('д/р', 'дл.р', 'дл/р', 'длинн', 'др', 'длру'):
                           {('повседневаная', 'пов', 'гол'):
                                ['Рубашка с длинным рукавом повседневная голубого цвета'],
                           ('пар', 'бел'):
                                ['Рубашка с длинным рукавом парадная белого цвета'],
                           },
                       ('к/р', 'кор.рук', 'кор.т.с.', 'кор.с.р', 'кор.р', 'к/рук', 'ско'):
                           {('повседневаная', 'пов', 'гол'):
                                ['Рубашка с коротким рукавом повседневная голубого цвета'],
                           ('пар', 'бел'):
                                ['Рубашка с коротким рукавом парадная белого цвета'],
                           },
                   },
                  ('брюки',):
                      {('лет',):
                           ['Брюки летние'],
                       ('дем'):
                            ['Брюки демисезонные'],
                      },
                  ('Джемпер', 'кардиган'): ['Джемпер трикотажный', 'Кардиган трикотажный'],
                  ('куртка',):
                    {
                        ('вет',): ['Куртка ветровка'],
                        ('зим',): ['Куртка зимняя']
                    },
                  ('вет',): ['Куртка ветровка'],
                ('Шапка', 'убор', 'голов', 'форменная'): ['головной убор зимний (трикотажная шапка)'],
                ('Жилет',):
                    {
                        ('утепленный', 'утеп', ): ['Жилет утепленный'],
                    },
                ('юбка',):
                    {
                        ('дем',): ['Юбка демисезонная'],
                        ('лет',): ['Юбка летняя'],
                    },
                ('платок',): ['Платок (шейный) или галстук-бант'],
              }


#Дефолтные типы одежды, если впервом словаре не нашлось то выставляется дефолтное значение по втором словарю
TYPE_CLOTHES_DEFAULT = {
        ('руб',):
               {('повседневаная', 'пов', 'гол'):
                    ['Рубашка с коротким рукавом повседневная голубого цвета'],
                ('пар', 'бел'):
                    ['Рубашка с коротким рукавом парадная белого цвета'],
                ('руб',):
                    ['Рубашка с коротким рукавом повседневная голубого цвета']
                },

}

def search_clothes(cloth, iter_clothes=TYPE_CLOTHES):
    for type_cloth in iter_clothes:
        for el in type_cloth:
            if el.lower() in cloth.replace(' ', '').lower():
                iter_clothes = iter_clothes[type_cloth]
                if type(iter_clothes) is list:
                    return iter_clothes
                else:
                    return search_clothes(cloth, iter_clothes)

def write_without_dict_clothes(tab_id, cloth):
    with open('Одежды который нет в словаре.txt', 'a') as f:
        if not re.findall(TYPE_CLOTHES_EXCLUDED, cloth.replace(' ', '').lower()):
            f.write(f'{tab_id} -- {cloth}\n')

def clothes(tabs):
    '''
    Получение данных от бухгалтерии
    :param tabs:
    :return:
    '''
    # добавление одежды которую пропустили
    df = pd.read_excel('data/бухКО.xlsx',
                       skiprows=2)
    tab_id = df['   Таб.№']
    cloths = df['Название основного средства']
    dates = df['Д/оприход.']
    cnt = df['                              Количество']
    for tab_el, cloth, date, c in zip(tab_id, cloths, dates, cnt):
        # проверяем только те табельные номера которые мы выбрали из файла с водителями
        if tab_el in tabs:
            re_clothes = search_clothes(cloth)
            if not re_clothes:
                re_clothes = search_clothes(cloth, TYPE_CLOTHES_DEFAULT)
                write_without_dict_clothes(tab_el, cloth)
            if re_clothes:
                for _ in range(c):
                    for re_cloth in re_clothes:
                        if tabs[tab_el]['clothes'].get(re_cloth):
                            tabs[tab_el]['clothes'][re_cloth]['date'].append(date.strftime('%d-%m-%Y'))
                        else:
                            tabs[tab_el]['clothes'][re_cloth] = {
                                'date': [date.strftime('%d-%m-%Y')],
                            }
            else:
                write_without_dict_clothes(tab_el, cloth)

    return tabs


def set_data_woman(tabs):
    cells = {
        'Рубашка с длинным рукавом повседневная голубого цвета': (12, 13),
        'Рубашка с длинным рукавом парадная белого цвета': [14],
        'Брюки демисезонные': [15],
        'Юбка демисезонная': [16],
        'Платок (шейный) или галстук-бант': [17],
        'Кардиган трикотажный': [18],
        'Жилет утепленный': [19],
        'Куртка ветровка': (20,),
        'Куртка зимняя': (21,),
        'головной убор зимний (трикотажная шапка)': [22],
        'Рубашка с коротким рукавом повседневная голубого цвета': (23, 24),
        'Рубашка с коротким рукавом парадная белого цвета': [25],
        'Брюки летние': (26, ),
        'Юбка летняя': (27, ),
        'Рубашка поло': (28, 29)
    }
    wb_obj = openpyxl.load_workbook(FILENAME)
    sheet_obj = wb_obj['Женский комплект']
    set_data(cells=cells, sheet=sheet_obj, tabs=tabs)
    wb_obj.save(FILENAME)

def set_data(cells, sheet, tabs):
    '''
        Создание excel по данным от бугалтерии
        :param data:
        :return:
        '''

    max_cell = max(cells.values(), key=lambda v: v[-1])[-1]
    min_cell = min(cells.values(), key=lambda v: v[0])[0]
    tab_coll = 1
    surname_coll_num = 4
    recr_coll_num = 11
    ogre_coll = 2

    i = 3
    for tab in tabs:
        for j in (range(min_cell, max_cell + 1)):
            cell_obj = sheet.cell(row=i, column=j)
            cell_obj.value = 'не выдано'

        '''
        Задается инфа по человек
        '''
        tab_id = sheet.cell(row=i, column=tab_coll)
        tab_id.value = tab

        surname = sheet.cell(row=i, column=surname_coll_num)
        surname.value = tabs[tab]['surname']

        recr = sheet.cell(row=i, column=recr_coll_num)
        recr.value = tabs[tab]['recr_date']

        ogre = sheet.cell(row=i, column=ogre_coll)
        ogre.value = tabs[tab]['ogre']

        '''
        Задается инфа по одежде
        '''
        for cloth in tabs[tab]['clothes']:
            date = tabs[tab]['clothes'][cloth]['date']
            if cells.get(cloth):
                # if нужен если в буглатерии дадут одежду которую не должны были довать,
                # Например юбку для мужчины
                for num_coll, date_value in zip(cells[cloth], date):
                    cell_cloth = sheet.cell(row=i, column=num_coll)
                    cell_cloth.value = date_value
        i += 1

def set_data_mans(tabs):

    cells = {
        'Рубашка с длинным рукавом повседневная голубого цвета': (12, 13),
        'Рубашка с длинным рукавом парадная белого цвета': [14],
        'Брюки демисезонные': [15],
        'Джемпер трикотажный': [16],
        'Жилет утепленный': [17],
        'Куртка ветровка': [18],
        'Куртка зимняя': [19],
        'головной убор зимний (трикотажная шапка)': (20,),
        'Рубашка с коротким рукавом повседневная голубого цвета': (21, 22),
        'Рубашка с коротким рукавом парадная белого цвета': [23],
        'Брюки летние': (24,),
        'Рубашка поло': [25, 26],
    }
    wb_obj = openpyxl.load_workbook('data/Шаблон Корпоративная одежда.xlsx')
    sheet_obj = wb_obj['Мужской комплект']
    set_data(cells=cells, sheet=sheet_obj, tabs=tabs)

    wb_obj.save(FILENAME)
    

def get_tab_id() -> dict:
    '''
    Получение табельных номеров для которых будет строится отчет
    '''
    SETTINGS = {
        'type_people': 'рег.гор.пасс.марш.'
    }
    wb_obj = openpyxl.load_workbook('data/водители.xlsx')
    sheet_obj = wb_obj.active
    start = 2
    tab_coll = 1
    surname_coll = 2
    sex_coll = 5
    recr_date_coll = 6
    type_people_coll = 4
    orge_coll = 3

    tabs = {}
    for tab_num in range(start, sheet_obj.max_row + 1):
        type_people = sheet_obj.cell(row=tab_num, column=type_people_coll).value
        if SETTINGS['type_people'] in type_people:
            id = int(sheet_obj.cell(row=tab_num, column=tab_coll).value)
            tabs[id] = {
                'sex': sheet_obj.cell(row=tab_num, column=sex_coll).value.upper(),
                'surname': sheet_obj.cell(row=tab_num, column=surname_coll).value,
                'recr_date': sheet_obj.cell(row=tab_num, column=recr_date_coll).value.strftime('%d-%m-%Y'),
                'ogre': sheet_obj.cell(row=tab_num, column=orge_coll).value,
                'clothes': {},
            }

    return tabs


def add_sizes():
    mans = ManSizes()   
    womans = WomanSizes()

    mans.get_data(file='data/Корпоративная одежда старая.xlsx')
    mans.set_data()

    womans.get_data(file='data/Корпоративная одежда старая.xlsx')
    womans.set_data()
try:
    os.remove('data/Корпоративная одежда ФЗ.xlsx')
    with open('Одежды который нет в словаре.txt', 'w') as f:
        f.write('')
except FileNotFoundError:
    pass
tabs = dict(sorted(get_tab_id().items()))

cloths_men = clothes({
    key: value for key, value in tabs.items() if value['sex'].upper() == 'мужской'.upper()
})
cloths_woman = clothes({
    key: value for key, value in tabs.items() if value['sex'].upper() == 'женский'.upper()
})

set_data_mans(cloths_men)
set_data_woman(cloths_woman)


add_sizes()



# try:
#     os.remove('data/Корпоративная одежда старая.xlsx')
# except FileNotFoundError:
#     pass
#TODO
'''
1. не корректное чтение файла бугалтерии
2. добавить один общий файл и перед зарузкой очищать его
'''
