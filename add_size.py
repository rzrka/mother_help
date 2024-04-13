import openpyxl


FILE = 'data/Корпоративная одежда ФЗ.xlsx'

GROUPS = {
    'data/Корпоративная одежда Зел.xlsx': ['Автоколонна 1', 'Автоколонна 2', 'Автоколонна 3', 'Автоколонна 7'],
    'data/Корпоративная одежда Левобережная.xlsx': ['Автоколонна 4', 'Автоколонна 5', 'Автоколонна 6'],
    'data/Корпоративная одежда Западный.xlsx': ['Автоколонна 8', 'Автоколонна 9', 'Автоколонна 10', 'Автоколонна 11'],

}


class Sizes:

    def __init__(self):
        self.start = 3
        self.tab_coll = None
        self.group_coll = 2
        self.height_coll = None
        self.breast_coll = None
        self.hips_coll = None
        self.waist_coll = None
        self.size_coll = None
        self.size_shoes = None
        self.data = {}
        self.sheet = None

    def get_data(self, file):
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj[self.sheet]
        self.data = {}
        for tab_num in range(self.start, sheet_obj.max_row + 1):
            # group = sheet_obj.cell(row=tab_num, column=self.group_coll).value
            # if group in GROUPS[file]:
            tab = sheet_obj.cell(row=tab_num, column=self.tab_coll).value
            height = sheet_obj.cell(row=tab_num, column=self.height_coll).value
            breast = sheet_obj.cell(row=tab_num, column=self.breast_coll).value
            hips = sheet_obj.cell(row=tab_num, column=self.hips_coll).value
            waist = sheet_obj.cell(row=tab_num, column=self.waist_coll).value
            size = sheet_obj.cell(row=tab_num, column=self.size_coll).value
            size_shoes = sheet_obj.cell(row=tab_num, column=self.size_shoes).value
            self.data[tab] = {
                'height': height if height else None,
                'breast': breast if breast else None,
                'hips': hips if hips else None,
                'waist': waist if waist else None,
                'size': size if size else None,
                'size_shoes': size_shoes if size_shoes else None,
                              }



    def set_data(self, file=FILE):
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj[self.sheet]
        for row in range(self.start, sheet_obj.max_row + 1):
            tab = sheet_obj.cell(row=row, column=self.tab_coll).value
            people = self.data.get(tab)
            if people:
                if people['height']:
                    height = sheet_obj.cell(row=row, column=self.height_coll)
                    height.value = people['height']

                if people['breast']:
                    breast = sheet_obj.cell(row=row, column=self.breast_coll)
                    breast.value = people['breast']

                if people['hips']:
                    hips = sheet_obj.cell(row=row, column=self.hips_coll)
                    hips.value = people['hips']

                if people['waist']:
                    waist = sheet_obj.cell(row=row, column=self.waist_coll)
                    waist.value = people['waist']

                if people['size']:
                    size = sheet_obj.cell(row=row, column=self.size_coll)
                    size.value = people['size']

                if people['size_shoes']:
                    size_shoes = sheet_obj.cell(row=row, column=self.size_shoes)
                    size_shoes.value = people['size_shoes']

        wb_obj.save(file)




class ManSizes(Sizes):

    def __init__(self):
        super().__init__()
        self.tab_coll = 1
        self.height_coll = 5
        self.breast_coll = 6
        self.hips_coll = 7
        self.waist_coll = 8
        self.size_coll = 9
        self.size_shoes = 10
        self.sheet = 'Мужской комплект'


class WomanSizes(Sizes):

    def __init__(self):
        super().__init__()
        self.tab_coll = 1
        self.height_coll = 5
        self.breast_coll = 6
        self.hips_coll = 7
        self.waist_coll = 8
        self.size_coll = 9
        self.size_shoes = 10
        self.sheet = 'Женский комплект'



# def sync(file, file_make):
#     diapason = {'start': 5, 'end': 9}
#     start = 6
#
#
#     wb = openpyxl.load_workbook(file)
#     wb_make = openpyxl.load_workbook(file_make)
#     for sheet_make in wb_make.worksheets:
#         ws = wb[sheet_make.title]
#         for row in range(start, sheet_make.max_row + 1):
#             for coll in range(diapason['start'], diapason['end']+1):
#                 value = sheet_make.cell(row=row, column=coll).value
#                 if value:
#                     cell = ws.cell(row=row, column=coll)
#                     cell.value = value
#     wb.save(file)
#
#
#
# sync(file=FILE, file_make='data/Корпоративная одежда Лев.xlsx')
# sync(file=FILE, file_make='data/Корпоративная одежда З.xlsx')
# sync(file=FILE, file_make='data/Корпоративная одежда Зап.xlsx')



